from telegram import Update, InlineKeyboardMarkup
from telegram.ext import ContextTypes, CallbackQueryHandler, ConversationHandler
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
from datetime import datetime
from custom_filters import PrivateChatAndAdmin, PermissionFilter
from common.keyboards import build_back_to_home_page_button, build_back_button
from common.lang_dicts import TEXTS, get_lang
from common.common import format_datetime
from admin.manage_users_settings.keyboards import build_manage_users_settings_keyboard
import tempfile
import os
import models


async def manage_users_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)
        keyboard = build_manage_users_settings_keyboard(lang)
        keyboard.append(build_back_button("back_to_admin_settings", lang=lang))
        keyboard.append(build_back_to_home_page_button(lang=lang, is_admin=True)[0])
        await update.callback_query.edit_message_text(
            text=TEXTS[lang]["manage_users_settings_title"],
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return ConversationHandler.END


manage_users_settings_handler = CallbackQueryHandler(
    manage_users_settings,
    "^manage_users_settings$|^back_to_manage_users_settings$",
)


async def export_users_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)

        await update.callback_query.answer(
            text=TEXTS[lang]["exporting_users"],
            show_alert=True,
        )

        await update.callback_query.delete_message()

        # إنشاء ملف Excel مؤقت
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            excel_path = tmp_file.name

            # إنشاء workbook جديد
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"

            # إضافة العناوين
            headers = [
                TEXTS[lang]["excel_user_id"],
                TEXTS[lang]["excel_username"],
                TEXTS[lang]["excel_name"],
                TEXTS[lang]["excel_language"],
                TEXTS[lang]["excel_is_admin"],
                TEXTS[lang]["excel_is_banned"],
                TEXTS[lang]["excel_created_at"],
            ]
            ws.append(headers)

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # جلب جميع المستخدمين من قاعدة البيانات
            with models.session_scope() as s:
                users = s.query(models.User).all()

                for user in users:
                    username = (
                        f"@{user.username}"
                        if user.username
                        else TEXTS[lang]["excel_no_username"]
                    )
                    lang_text = (
                        TEXTS[lang][f"lang_{user.lang.name.lower()}"]
                        if user.lang
                        else TEXTS[lang]["excel_unknown"]
                    )
                    is_admin = (
                        TEXTS[lang]["excel_yes"]
                        if user.is_admin
                        else TEXTS[lang]["excel_no"]
                    )
                    is_banned = (
                        TEXTS[lang]["excel_yes"]
                        if user.is_banned
                        else TEXTS[lang]["excel_no"]
                    )
                    created_at = (
                        format_datetime(user.created_at)
                        if user.created_at
                        else TEXTS[lang]["excel_unknown"]
                    )

                    ws.append(
                        [
                            user.user_id,
                            username,
                            user.name,
                            lang_text,
                            is_admin,
                            is_banned,
                            created_at,
                        ]
                    )

            # ضبط عرض الأعمدة
            ws.column_dimensions["A"].width = 15  # User ID
            ws.column_dimensions["B"].width = 20  # Username
            ws.column_dimensions["C"].width = 25  # Name
            ws.column_dimensions["D"].width = 15  # Language
            ws.column_dimensions["E"].width = 12  # Is Admin
            ws.column_dimensions["F"].width = 12  # Is Banned
            ws.column_dimensions["G"].width = 20  # Created At

            # حفظ الملف
            wb.save(excel_path)

        try:
            # إرسال الملف
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.xlsx"

            with open(excel_path, "rb") as excel_file:
                await context.bot.send_document(
                    chat_id=update.effective_user.id,
                    document=excel_file,
                    filename=filename,
                )

            text = TEXTS[lang]["users_exported_success"]
        except Exception as e:
            text = TEXTS[lang]["export_error"]

        # حذف الملف المؤقت
        if os.path.exists(excel_path):
            os.unlink(excel_path)

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=text + "\n\n" + TEXTS[lang]["continue_with_admin_command"],
        )


export_users_handler = CallbackQueryHandler(
    export_users_to_excel,
    "^export_users_to_excel$",
)
